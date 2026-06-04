from datetime import timedelta
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

DAILY_PROMO_TEXT = "PRIME BURGERS/ITALIAN BROWNIE SAMPLE/TOAST/MISC"

# ── Colors matching screenshot ────────────────────────────────────────────────
HEADER_FILL     = PatternFill("solid", fgColor="FFFFFF")   # white
ROLE_FILL       = PatternFill("solid", fgColor="D6DCE4")   # light blue/gray
PROMO_FILL      = PatternFill("solid", fgColor="F8CBAD")   # peach/orange
ALT_FILL        = PatternFill("solid", fgColor="FFFFFF")   # no alternating gray
WHITE_FILL      = PatternFill("solid", fgColor="FFFFFF")
PRIVATE_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green

# ── Fonts ─────────────────────────────────────────────────────────────────────
HEADER_FONT     = Font(name="Arial", bold=False, color="000000", size=10)
ROLE_FONT       = Font(name="Arial", bold=True, color="000000", size=10)
PROMO_FONT      = Font(name="Arial", italic=True, color="000000", size=10)
BODY_FONT       = Font(name="Arial", size=10)

# ── Alignment ─────────────────────────────────────────────────────────────────
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Border ────────────────────────────────────────────────────────────────────
THIN        = Side(style="thin", color="000000")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROLE_ORDER = ["Manager", "Host", "Bar", "Server", "Assistant", "Dish Washer"]


def _fmt_arrival(dt):
    local = timezone.localtime(dt)
    hour, minute = local.hour, local.minute
    period = "AM" if hour < 12 else "PM"
    hour_12 = hour % 12 or 12
    return f"{hour_12}:{minute:02d}{period}" if minute else f"{hour_12}{period}"


def _collect_schedule_data(schedule_week, department=None):
    from .models import Assignment

    week_start = schedule_week.week_start
    week_end   = week_start + timedelta(days=6)

    qs = (
        Assignment.objects
        .select_related("employee__primary_role", "shift__role")
        .filter(
            shift__start_time__date__gte=week_start,
            shift__start_time__date__lte=week_end,
        )
        .order_by("shift__start_time", "employee__name")
    )

    if department:
        qs = qs.filter(employee__primary_role__department=department)

    employee_map = {}
    for assignment in qs:
        emp        = assignment.employee
        shift      = assignment.shift
        shift_date = timezone.localtime(shift.start_time).date()
        day_index  = (shift_date - week_start).days
        if day_index < 0 or day_index > 6:
            continue
        if emp.id not in employee_map:
            employee_map[emp.id] = {
                "employee":     emp,
                "primary_role": emp.primary_role,
                "days":         {},
            }
        employee_map[emp.id]["days"][day_index] = _fmt_arrival(shift.start_time)

    role_groups = {}
    unroled     = []
    for ed in employee_map.values():
        role = ed["primary_role"]
        if role:
            if role.id not in role_groups:
                role_groups[role.id] = {"role": role, "employees": []}
            role_groups[role.id]["employees"].append(ed)
        else:
            unroled.append(ed)

    def role_sort_key(g):
        try:
            return ROLE_ORDER.index(g["role"].name)
        except ValueError:
            return len(ROLE_ORDER)

    sorted_groups = sorted(role_groups.values(), key=role_sort_key)
    for g in sorted_groups:
        g["employees"].sort(key=lambda e: e["employee"].name)

    if unroled:
        unroled.sort(key=lambda e: e["employee"].name)
        sorted_groups.append({"role": None, "employees": unroled})

    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    return sorted_groups, week_dates


def _apply_cell(ws, row, col, value, font, fill, alignment, border=CELL_BORDER):
    cell            = ws.cell(row=row, column=col, value=value)
    cell.font       = font
    cell.fill       = fill
    cell.alignment  = alignment
    cell.border     = border
    return cell


def _write_full_row(ws, row, value, font, fill, alignment, row_height=16):
    """Write a value in col 1 and fill cols 2-8 with the same style."""
    _apply_cell(ws, row, 1, value, font, fill, alignment)
    for col in range(2, 9):
        _apply_cell(ws, row, col, "", font, fill, CENTER)
    ws.row_dimensions[row].height = row_height


def generate_schedule_excel(schedule_week, department=None):
    sorted_groups, week_dates = _collect_schedule_data(schedule_week, department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Column widths
    ws.column_dimensions["A"].width = 18
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i + 1)].width = 15

    # ── Row 1: day/date header ─────────────────────────────────────────────
    _apply_cell(ws, 1, 1, "160 Main", HEADER_FONT, HEADER_FILL, CENTER)
    for i, date in enumerate(week_dates):
        label = f"{DAY_NAMES[i]}\n{date.strftime('%-m/%-d/%Y')}"
        _apply_cell(ws, 1, i + 2, label, HEADER_FONT, HEADER_FILL, CENTER)

    ws.row_dimensions[1].height = 34

    current_row = 2

    # ── Static PRIVATE PARTY block: 2 green rows ──────────────────────────
    _write_full_row(ws, current_row,     "PRIVATE PARTY", ROLE_FONT, PRIVATE_FILL, LEFT, 18)
    _write_full_row(ws, current_row + 1, "",              ROLE_FONT, PRIVATE_FILL, LEFT, 18)
    current_row += 2

    # ── Static DAILY PROMO row ─────────────────────────────────────────────
    _write_full_row(ws, current_row, DAILY_PROMO_TEXT, PROMO_FONT, PROMO_FILL, LEFT, 18)
    current_row += 1

    # ── Role groups with employees ─────────────────────────────────────────
    for group in sorted_groups:
        role      = group["role"]
        role_name = role.name.upper() if role else "OTHER"

        _write_full_row(ws, current_row, role_name, ROLE_FONT, ROLE_FILL, LEFT, 18)
        current_row += 1

        for idx, ed in enumerate(group["employees"]):
            emp  = ed["employee"]
            days = ed["days"]
            fill = WHITE_FILL

            _apply_cell(ws, current_row, 1, emp.name, BODY_FONT, fill, LEFT)
            for d in range(7):
                _apply_cell(ws, current_row, d + 2, days.get(d, ""), BODY_FONT, fill, CENTER)

            ws.row_dimensions[current_row].height = 17
            current_row += 1

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()